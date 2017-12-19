from libnmap.parser import NmapParser
from libnmap.process import NmapProcess
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import manuf

# start a new nmap scan on localhost with some specific options
def do_scan(targets, options):
    parsed = None
    nmproc = NmapProcess(targets, options)
    rc = nmproc.run()
    if rc != 0:
        print("nmap scan failed: {0}".format(nmproc.stderr))
    print(type(nmproc.stdout))

    try:
        parsed = NmapParser.parse(nmproc.stdout)
    except NmapParserException as e:
        print("Exception raised while parsing scan: {0}".format(e.msg))

    return parsed


# export scan results from a nmap report
def print_scan(nmap_report,fileName,iPRange,hostNet):
    print("Starting Nmap {0} ( http://nmap.org ) at {1}".format(
        nmap_report.version,
        nmap_report.started))
    
    #initialize the Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Nmap Scan"
    ws['A1'] = "IP"
    ws['B1'] = "Hostname"
    ws['C1'] = "Port/Protocol/Service"
    ws['D1'] = "OS"
    ws['E1'] = "Scan Vendor"
    ws['F1'] = "MAC"
    ws['G1'] = "MAC Vendor"
    ws['H1'] = "Priority/Risk"
    ws['I1'] = "Comments"

    #highlight the top cells
    for j in range(1,10):
        fillColor = PatternFill("solid", fgColor="2672EC")
        cellStyle = ws.cell(row=1, column=j)
        cellStyle.fill = fillColor
    
    #Write the IP Addresses and highlight the cells in our DHCP range
    for i in range(iPRange):
        ws.cell(row=(i+2), column=1, value=hostNet+str(i))
        if i in range(100, 200):
            fill = PatternFill("solid", fgColor="e0a4ea")
            for j in range(1,10):
                cellStyle = ws.cell(row=(i+2), column=j)
                cellStyle.fill = fill

    #Walk through the hosts found
    for host in nmap_report.hosts:
        #Set the row to be in line with the Correct Address of scanned host
        for i in range(iPRange):
            if hostNet+str(i) == host.address:
                row = i + 2
                break
        #This is to filter out the IP that do not have a host
        #It should work that if there are any open ports its True and continues or if there was found a MAC address its true
        if host.get_open_ports() or host.mac != '':
            for j in range(1,10):
                fillColor = PatternFill("solid", fgColor="FFFFFF")
                cellStyle = ws.cell(row=1, column=j)
                cellStyle.fill = fillColor
            
            #If the host has a hostname add it
            if len(host.hostnames):
                tmp_host = ','.join(host.hostnames)
            else:
                tmp_host = 'None set'

            #Add the hostname if one exists
            ws.cell(row=row, column=2, value=tmp_host)
            
            #Creates a list of the open ports and services, then converts to string to be able to track for security
            portList = []
            for serv in host.services:
                pserv = "{0:>5s}/{1:3s}/{2}".format(
                        str(serv.port),
                        serv.protocol,
                        serv.service)
                portList.append(pserv)
            portString = ','.join(portList)
            ws.cell(row=row, column=3, value=portString)

            #Adds the OS if it was able to be found
            if len(host.os_class_probabilities()):
                os = host.os_class_probabilities()
                osEle = str(os[0])
                osList = osEle.split("\n")
                osPos = str(osList[0])
                ws.cell(row=row, column=4, value=osPos)
            else:
                ws.cell(row=row, column=4, value="Unable to detect")

            #If the vendor could be determined by the NMAP adds it
            venCheck = host.vendor
            if not venCheck:
                ws.cell(row=row, column=5, value=str(host.vendor))
            else:
                ws.cell(row=row, column=5, value="Unable to detect")
            
            #Uses manuf to find the Manufacturer based upon the MAC
            if host.mac != '':
                ws.cell(row=row, column=6, value=str(host.mac))
                macVen = manuf.MacParser()
                ws.cell(row=row, column=7, value=macVen.get_manuf(host.mac))
            else:
                ws.cell(row=row, column=6, value="No MAC")
                ws.cell(row=row, column=7, value="No MAC")

        #If there was no Host found on the IP, then it posts that the Address is open and highlights them
        else:
            ws.cell(row=row, column=2, value='Open Addr')
            for j in range(1,10):
                fillColor = PatternFill("solid", fgColor="AA9CD2")
                cellStyle = ws.cell(row=1, column=j)
                cellStyle.fill = fillColor

    wb.save(fileName)
    print(nmap_report.summary)


if __name__ == "__main__":
    fileName = "scan.xlsx"
    target = "192.168.10.0/24"
    hostNet = "192.168.10."
    iPRange = 256
    ##Was working on a way for a user to input an IP range, found it easier to just hard code it for internal use
    # slash = "/"
    # subDet = target.find(slash)
    # if subDet == -1:
    #     subNet = 32
    #     hostNet = target
    # else:
    #     subSplit = target.split("/")
    #     subNet = subDet[1]
    #     targetNet = target[0]
    #     targetList = targetNet.split(".")
    #     targetString = str(targetList[-2:])
    # iPRange = 2 ** (32 - subNet)
    print("Beginning Scan")
    report = do_scan(target, "-A")
    print("Reporting")
    if report:
        print_scan(report,fileName,iPRange,hostNet)
    else:
        print("No results returned")
